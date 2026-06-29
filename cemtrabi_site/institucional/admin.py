from django.contrib import admin
from django.contrib.auth.admin import GroupAdmin as BaseGroupAdmin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.auth.models import Group, User
from django.http import HttpResponse
from django.utils import timezone
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from unfold.admin import ModelAdmin
from unfold.forms import AdminPasswordChangeForm, UserChangeForm, UserCreationForm

from .models import Encaminhamento, Lead


admin.site.site_header = 'CEMTRABI'
admin.site.site_title = 'CEMTRABI Admin'
admin.site.index_title = 'Painel operacional'


try:
    admin.site.unregister(User)
except admin.sites.NotRegistered:
    pass

try:
    admin.site.unregister(Group)
except admin.sites.NotRegistered:
    pass


@admin.register(User)
class UserAdmin(BaseUserAdmin, ModelAdmin):
    form = UserChangeForm
    add_form = UserCreationForm
    change_password_form = AdminPasswordChangeForm
    compressed_fields = True


@admin.register(Group)
class GroupAdmin(BaseGroupAdmin, ModelAdmin):
    compressed_fields = True


def exportar_para_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Leads CEMTRABI'

    headers = [
        'Empresa',
        'CNPJ',
        'Responsável',
        'E-mail',
        'Telefone',
        'Serviço',
        'Status',
        'Origem',
        'Mensagem',
        'Data de Criação',
    ]
    ws.append(headers)

    for lead in queryset:
        ws.append([
            lead.empresa,
            lead.cnpj,
            lead.responsavel,
            lead.email,
            lead.telefone,
            lead.get_servico_display(),
            lead.get_status_display(),
            lead.origem,
            lead.mensagem,
            timezone.localtime(lead.criado_em).strftime('%d/%m/%Y %H:%M'),
        ])

    aplicar_estilo_planilha(ws)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=leads_cemtrabi.xlsx'

    wb.save(response)
    return response


exportar_para_excel.short_description = 'Exportar leads selecionados para Excel'


def listar_exames_encaminhamento(encaminhamento):
    exames = [
        ('Exame Clínico', encaminhamento.exame_clinico),
        ('Audiometria Tonal', encaminhamento.audiometria),
        ('Acuidade Visual', encaminhamento.acuidade_visual),
        ('Eletrocardiograma', encaminhamento.eletrocardiograma),
        ('Eletroencefalograma', encaminhamento.eletroencefalograma),
        ('Espirometria', encaminhamento.espirometria),
        ('Avaliação Psicossocial', encaminhamento.avaliacao_psicossocial),
        ('Raio X Tórax Padrão OIT', encaminhamento.raio_x_torax),
        ('Teste de Equilíbrio/Romberg', encaminhamento.teste_romberg),
        ('Hemograma Completo com Plaquetas', encaminhamento.hemograma),
        ('Glicemia em Jejum', encaminhamento.glicemia),
        ('Ácido Hipúrico', encaminhamento.acido_hipurico),
        ('Ácido Metil Hipúrico', encaminhamento.acido_metil_hipurico),
        ('Grupo Sanguíneo/Fator RH', encaminhamento.grupo_sanguineo_fator_rh),
    ]

    exames_selecionados = [nome for nome, marcado in exames if marcado]

    if encaminhamento.outro_exame and encaminhamento.descricao_outro_exame:
        exames_selecionados.append(encaminhamento.descricao_outro_exame)

    if encaminhamento.outro_exame and not encaminhamento.descricao_outro_exame:
        exames_selecionados.append('Outro exame')

    return '\n'.join(exames_selecionados)


def aplicar_estilo_planilha(ws):
    header_fill = PatternFill(fill_type='solid', fgColor='D9EAD3')
    header_font = Font(bold=True, color='183B34')
    header_alignment = Alignment(horizontal='center', vertical='center')
    body_alignment = Alignment(vertical='top', wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row in ws.iter_rows(min_row=2):
        max_lines = 1
        for cell in row:
            cell.alignment = body_alignment
            if isinstance(cell.value, str):
                max_lines = max(max_lines, cell.value.count('\n') + 1)
        ws.row_dimensions[row[0].row].height = max(18, max_lines * 15)

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value:
                cell_lines = str(cell.value).split('\n')
                max_length = max(max_length, *(len(line) for line in cell_lines))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 45)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def exportar_encaminhamentos_para_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Encaminhamentos'

    headers = [
        'Nome',
        'Empresa',
        'CNPJ',
        'Função',
        'Setor',
        'Tipo de Exame',
        'Exames Selecionados',
        'Telefone',
        'Email',
        'Nome do Emitente',
        'Data do Encaminhamento',
        'Data de Criação',
    ]
    ws.append(headers)

    for encaminhamento in queryset:
        ws.append([
            encaminhamento.nome,
            encaminhamento.empresa,
            encaminhamento.cnpj,
            encaminhamento.funcao,
            encaminhamento.setor,
            encaminhamento.get_tipo_exame_display(),
            listar_exames_encaminhamento(encaminhamento),
            encaminhamento.telefone,
            encaminhamento.email,
            encaminhamento.nome_emitente,
            encaminhamento.data_encaminhamento.strftime('%d/%m/%Y'),
            timezone.localtime(encaminhamento.created_at).strftime('%d/%m/%Y %H:%M'),
        ])

    aplicar_estilo_planilha(ws)

    data_arquivo = timezone.localdate().strftime('%Y-%m-%d')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename=encaminhamentos_cemtrabi_{data_arquivo}.xlsx'
    )

    wb.save(response)
    return response


exportar_encaminhamentos_para_excel.short_description = (
    'Exportar encaminhamentos selecionados para Excel'
)


def badge(texto, cor='teal'):
    return format_html(
        '<span class="cem-unfold-badge cem-unfold-badge-{}">{}</span>',
        cor,
        texto,
    )


@admin.register(Lead)
class LeadAdmin(ModelAdmin):
    compressed_fields = True
    warn_unsaved_form = True
    list_fullwidth = True
    list_filter_sheet = True
    list_filter_submit = True
    list_display = (
        'empresa_card',
        'responsavel',
        'email',
        'telefone',
        'servico_badge',
        'status_badge',
        'criado_em_formatado',
    )
    list_filter = ('servico', 'status', 'criado_em')
    search_fields = ('empresa', 'cnpj', 'responsavel', 'email', 'telefone')
    readonly_fields = ('criado_em', 'origem')
    date_hierarchy = 'criado_em'
    ordering = ('-criado_em',)
    list_per_page = 25
    actions = [exportar_para_excel]

    fieldsets = (
        ('Empresa', {
            'fields': ('empresa', 'cnpj', 'servico'),
        }),
        ('Responsável', {
            'fields': ('responsavel', 'email', 'telefone'),
        }),
        ('Mensagem', {
            'fields': ('mensagem',),
        }),
        ('Controle interno', {
            'classes': ('collapse',),
            'fields': ('status', 'origem', 'criado_em'),
        }),
    )

    def empresa_card(self, obj):
        return format_html(
            '<strong class="cem-unfold-title">{}</strong><br>'
            '<span class="cem-unfold-muted">{}</span>',
            obj.empresa,
            obj.cnpj,
        )

    empresa_card.short_description = 'Empresa'
    empresa_card.admin_order_field = 'empresa'

    def servico_badge(self, obj):
        cores = {
            'medicina': 'teal',
            'seguranca': 'cyan',
            'treinamentos': 'mint',
        }
        return badge(obj.get_servico_display(), cores.get(obj.servico, 'teal'))

    servico_badge.short_description = 'Serviço'
    servico_badge.admin_order_field = 'servico'

    def status_badge(self, obj):
        cores = {
            'novo': 'blue',
            'contato': 'cyan',
            'orcamento': 'amber',
            'fechado': 'green',
        }
        return badge(obj.get_status_display(), cores.get(obj.status, 'teal'))

    status_badge.short_description = 'Status'
    status_badge.admin_order_field = 'status'

    def criado_em_formatado(self, obj):
        return timezone.localtime(obj.criado_em).strftime('%d/%m/%Y %H:%M')

    criado_em_formatado.short_description = 'Criado em'
    criado_em_formatado.admin_order_field = 'criado_em'


@admin.register(Encaminhamento)
class EncaminhamentoAdmin(ModelAdmin):
    compressed_fields = True
    warn_unsaved_form = True
    list_fullwidth = True
    list_filter_sheet = True
    list_filter_submit = True
    list_horizontal_scrollbar_top = True
    list_display = (
        'colaborador_card',
        'empresa_card',
        'tipo_exame_badge',
        'status_badge',
        'email_status',
        'criado_em',
        'link_docx',
        'link_pcmso',
    )
    list_filter = ('tipo_exame', 'status', 'email_enviado', 'created_at', 'empresa')
    search_fields = ('nome', 'cpf', 'rg', 'empresa', 'cnpj', 'funcao', 'setor')
    readonly_fields = (
        'idade',
        'email_enviado',
        'link_docx',
        'link_pcmso',
        'created_at',
    )
    date_hierarchy = 'created_at'
    ordering = ('-created_at',)
    list_per_page = 25
    save_on_top = True
    actions = [exportar_encaminhamentos_para_excel]

    fieldsets = (
        ('Dados do Colaborador', {
            'fields': (
                ('nome', 'cpf'),
                ('rg', 'data_nascimento', 'idade'),
                ('telefone', 'email'),
            )
        }),
        ('Dados da Empresa', {
            'fields': (
                ('empresa', 'cnpj'),
                'endereco_empresa',
                ('nome_emitente', 'data_encaminhamento'),
            )
        }),
        ('Função e Exposição Ocupacional', {
            'fields': (
                ('funcao', 'setor'),
                ('data_admissao', 'tipo_exame'),
                'riscos_ocupacionais',
            )
        }),
        ('Exames Complementares', {
            'classes': ('collapse',),
            'fields': (
                ('exame_clinico', 'audiometria', 'hemograma'),
                ('glicemia', 'acuidade_visual', 'eletrocardiograma'),
                ('eletroencefalograma', 'espirometria', 'avaliacao_psicossocial'),
                ('raio_x_torax', 'teste_romberg'),
                ('acido_hipurico', 'acido_metil_hipurico'),
                ('grupo_sanguineo_fator_rh', 'outro_exame'),
                'descricao_outro_exame',
            )
        }),
        ('Arquivos', {
            'fields': (
                ('pcmso', 'link_pcmso'),
                ('docx_gerado', 'link_docx'),
            )
        }),
        ('Controle Interno', {
            'classes': ('collapse',),
            'fields': (
                ('status', 'email_enviado'),
                'created_at',
            )
        }),
    )

    def colaborador_card(self, obj):
        return format_html(
            '<strong class="cem-unfold-title">{}</strong><br>'
            '<span class="cem-unfold-muted">{} - {}</span>',
            obj.nome,
            obj.cpf,
            obj.funcao,
        )

    colaborador_card.short_description = 'Colaborador'
    colaborador_card.admin_order_field = 'nome'

    def empresa_card(self, obj):
        return format_html(
            '<strong class="cem-unfold-title">{}</strong><br>'
            '<span class="cem-unfold-muted">{}</span>',
            obj.empresa,
            obj.setor,
        )

    empresa_card.short_description = 'Empresa'
    empresa_card.admin_order_field = 'empresa'

    def tipo_exame_badge(self, obj):
        return badge(obj.get_tipo_exame_display(), 'mint')

    tipo_exame_badge.short_description = 'Tipo'
    tipo_exame_badge.admin_order_field = 'tipo_exame'

    def status_badge(self, obj):
        cores = {
            'pendente': 'slate',
            'recebido': 'blue',
            'agendado': 'amber',
            'concluido': 'green',
            'cancelado': 'red',
        }
        return badge(obj.get_status_display(), cores.get(obj.status, 'teal'))

    status_badge.short_description = 'Status'
    status_badge.admin_order_field = 'status'

    def email_status(self, obj):
        if obj.email_enviado:
            return badge('Enviado', 'green')
        return badge('Pendente', 'amber')

    email_status.short_description = 'E-mail'
    email_status.admin_order_field = 'email_enviado'

    def criado_em(self, obj):
        return timezone.localtime(obj.created_at).strftime('%d/%m/%Y %H:%M')

    criado_em.short_description = 'Criado em'
    criado_em.admin_order_field = 'created_at'

    def link_docx(self, obj):
        if not obj or not obj.docx_gerado:
            return mark_safe('<span class="cem-unfold-muted">-</span>')
        return format_html(
            '<a class="cem-unfold-link" href="{}" target="_blank" rel="noopener">DOCX</a>',
            obj.docx_gerado.url,
        )

    link_docx.short_description = 'DOCX'

    def link_pcmso(self, obj):
        if not obj or not obj.pcmso:
            return mark_safe('<span class="cem-unfold-muted">-</span>')
        return format_html(
            '<a class="cem-unfold-link" href="{}" target="_blank" rel="noopener">PCMSO</a>',
            obj.pcmso.url,
        )

    link_pcmso.short_description = 'PCMSO'
