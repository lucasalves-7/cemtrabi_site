from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Lead


def exportar_para_excel(modeladmin, request, queryset):
    """
    Exporta os leads selecionados para Excel (.xlsx)
    com ajuste automático de largura das colunas.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads CEMTRABI"

    # Cabeçalhos
    headers = [
        'Empresa',
        'CNPJ',
        'Responsável',
        'E-mail',
        'Telefone',
        'Serviço',
        'Status',
        'Origem',
        'Mensagem',
        'Data de Criação',
    ]

    ws.append(headers)

    # Dados
    for lead in queryset:
        ws.append([
            lead.empresa,
            lead.cnpj,
            lead.responsavel,
            lead.email,
            lead.telefone,
            lead.get_servico_display(),
            lead.get_status_display(),
            lead.origem,
            lead.mensagem,
            lead.criado_em.strftime('%d/%m/%Y %H:%M'),
        ])

    # 🔹 AUTO AJUSTE DE COLUNAS
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 2

    # Resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=leads_cemtrabi.xlsx'

    wb.save(response)
    return response


@admin.register(Lead)
class LeadAdmin(admin.ModelAdmin):
    list_display = (
        'empresa',
        'responsavel',
        'email',
        'telefone',
        'servico',
        'status',
        'criado_em',
    )

    list_filter = ('servico', 'status')
    search_fields = ('empresa', 'responsavel', 'email')

    actions = [exportar_para_excel]
    exportar_para_excel.short_description = "Exportar leads selecionados para Excel"
